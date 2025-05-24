"""
Excel file processor with memory-efficient sheet-by-sheet processing.
"""
import os
import time
import re
from typing import Dict, Generator, List, Optional, Any, Union, Tuple

import pandas as pd
import numpy as np
from tqdm import tqdm
import openpyxl

from app.utils.logging_config import get_logger
from app.processors.base_processor import BaseProcessor

logger = get_logger(__name__)

class ExcelProcessor(BaseProcessor):
    """
    Process large Excel files sheet by sheet with memory optimization.
    
    This class handles Excel files efficiently by:
    1. Processing sheets individually to minimize memory usage
    2. Reading data in chunks within each sheet
    3. Optimizing data types to reduce memory footprint
    """
    
    def __init__(self, **kwargs):
        """Initialize the Excel processor with shared parameters."""
        super().__init__(**kwargs)
        # Excel-specific settings
        self.sheet_name = kwargs.get('sheet_name', None)  # None means all sheets
        self.header = kwargs.get('header', 0)
        self.skiprows = kwargs.get('skiprows', None)
        self.usecols = kwargs.get('usecols', None)
        self.dtype = kwargs.get('dtype', None)
        self.date_format = kwargs.get('date_format', None)
        self.date_columns = kwargs.get('date_columns', [])
        self.skip_blank_lines = kwargs.get('skip_blank_lines', True)
        
    def read_file(self, file_path: str) -> Generator[pd.DataFrame, None, None]:
        """
        Read an Excel file in memory-efficient chunks by sheet.
        
        Args:
            file_path: Path to the Excel file
            
        Yields:
            DataFrame chunks from the Excel file
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
            
        # Get file info
        file_info = self.get_file_info(file_path)
        
        # Determine sheets to process
        sheets_to_process = []
        if self.sheet_name is None:
            # Process all sheets
            sheets_to_process = file_info.get('sheet_names', [])
        elif isinstance(self.sheet_name, (list, tuple)):
            # Process specified sheets
            sheets_to_process = self.sheet_name
        else:
            # Process single sheet
            sheets_to_process = [self.sheet_name]
            
        # Determine optimal chunk size if not provided
        chunk_size = self.get_optimal_chunk_size(file_path)
        if chunk_size is None:
            chunk_size = 1000  # Default fallback for Excel
            logger.info(f"Using default chunk size: {chunk_size} rows")
        else:
            # Convert MB to approximate rows (assuming ~200 bytes per row for Excel)
            chunk_size = int(chunk_size * 1024 * 1024 / 200)
            logger.info(f"Auto-calculated chunk size: {chunk_size} rows")
            
        # Create a progress bar for sheets if showing progress
        sheet_pbar = tqdm(sheets_to_process, desc=f"Processing sheets in {os.path.basename(file_path)}", 
                         disable=not self.show_progress)
        
        # Process each sheet
        for sheet_name in sheet_pbar:
            try:
                # Update progress description
                if self.show_progress:
                    sheet_pbar.set_description(f"Processing sheet: {sheet_name}")
                
                # Get row count if available
                sheet_info = file_info.get('sheets', {}).get(sheet_name, {})
                total_rows = sheet_info.get('row_count', 0)
                
                # If the sheet is small enough, read it all at once
                if total_rows > 0 and total_rows <= chunk_size:
                    try:
                        # Read the entire sheet
                        df = pd.read_excel(
                            file_path,
                            sheet_name=sheet_name,
                            header=self.header,
                            skiprows=self.skiprows,
                            usecols=self.usecols,
                            dtype=self.dtype
                        )
                        
                        # Add sheet name as metadata
                        df['_sheet_name'] = sheet_name
                        
                        # Optimize data types
                        try:
                            # Use base class method to detect optimal types
                            optimal_types = self.detect_data_type(df)
                            # Apply the types
                            for col, dtype in optimal_types.items():
                                if col in df.columns:
                                    try:
                                        df[col] = df[col].astype(dtype)
                                    except:
                                        pass  # Skip if conversion fails
                        except:
                            pass  # Use original if optimization fails
                        
                        # Handle date columns
                        df = self._process_date_columns(df)
                        
                        yield df
                        
                    except Exception as e:
                        logger.error(f"Error reading sheet {sheet_name}: {str(e)}")
                        # Create error DataFrame
                        error_df = pd.DataFrame({
                            'sheet_name': [sheet_name],
                            'error': [str(e)]
                        })
                        yield error_df
                        
                    # Clean up memory
                    self.memory_monitor.cleanup_memory()
                    continue
                
                # For larger sheets, read in chunks
                # Manually iterating through rows using openpyxl for better memory management
                workbook = openpyxl.load_workbook(
                    filename=file_path, 
                    read_only=True, 
                    data_only=True
                )
                
                try:
                    sheet = workbook[sheet_name]
                    
                    # Get max row/column counts
                    rows = sheet.max_row
                    cols = sheet.max_column
                    
                    # Determine header row
                    header_row = self.header
                    
                    # Get header values if header exists
                    if header_row is not None:
                        header_values = []
                        for row in sheet.iter_rows(min_row=header_row+1, max_row=header_row+1):
                            header_values = [cell.value for cell in row]
                            break
                    else:
                        # Generate default column names
                        header_values = [f"Column_{i}" for i in range(cols)]
                    
                    # Calculate starting row (accounting for header and skiprows)
                    start_row = 0
                    if header_row is not None:
                        start_row = header_row + 1
                    if self.skiprows is not None:
                        if isinstance(self.skiprows, int):
                            start_row += self.skiprows
                        else:
                            # Assume it's a list of row indices
                            start_row = max(start_row, max(self.skiprows) + 1)
                    
                    # Process in chunks
                    chunk_data = []
                    row_count = 0
                    
                    # Create progress bar for rows if total known
                    row_range = range(start_row + 1, rows + 1)
                    row_pbar = tqdm(total=len(row_range), desc=f"Reading rows", 
                                  disable=not self.show_progress)
                    
                    # Iterate through rows in chunks
                    for i, row in enumerate(sheet.iter_rows(min_row=start_row + 1, values_only=True)):
                        # Skip rows if they are in skiprows
                        if self.skiprows is not None and isinstance(self.skiprows, list) and i in self.skiprows:
                            continue
                        
                        # Skip blank rows if specified
                        if self.skip_blank_lines and all(cell is None or str(cell).strip() == "" for cell in row):
                            continue
                            
                        # Filter columns if usecols specified
                        if self.usecols is not None:
                            if isinstance(self.usecols, list):
                                # List of column indices
                                row = [row[i] if i < len(row) else None for i in self.usecols]
                            elif callable(self.usecols):
                                # Function to filter columns
                                row = [val for i, val in enumerate(row) if self.usecols(i)]
                        
                        chunk_data.append(row)
                        row_count += 1
                        
                        # Update row progress
                        if self.show_progress:
                            row_pbar.update(1)
                        
                        # If we've reached chunk size, yield data
                        if len(chunk_data) >= chunk_size:
                            # Create DataFrame from chunk
                            df = pd.DataFrame(chunk_data, columns=header_values)
                            
                            # Add sheet name as metadata
                            df['_sheet_name'] = sheet_name
                            
                            # Optimize data types
                            try:
                                # Use base class method to detect optimal types
                                optimal_types = self.detect_data_type(df)
                                # Apply the types
                                for col, dtype in optimal_types.items():
                                    if col in df.columns:
                                        try:
                                            df[col] = df[col].astype(dtype)
                                        except:
                                            pass  # Skip if conversion fails
                            except:
                                pass  # Use original if optimization fails
                            
                            # Handle date columns
                            df = self._process_date_columns(df)
                            
                            yield df
                            
                            # Reset chunk data
                            chunk_data = []
                            
                            # Clean up memory
                            self.memory_monitor.cleanup_memory()
                    
                    # Close row progress bar
                    if self.show_progress:
                        row_pbar.close()
                    
                    # Process final chunk if any data remains
                    if chunk_data:
                        df = pd.DataFrame(chunk_data, columns=header_values)
                        
                        # Add sheet name as metadata
                        df['_sheet_name'] = sheet_name
                        
                        # Optimize data types
                        try:
                            # Use base class method to detect optimal types
                            optimal_types = self.detect_data_type(df)
                            # Apply the types
                            for col, dtype in optimal_types.items():
                                if col in df.columns:
                                    try:
                                        df[col] = df[col].astype(dtype)
                                    except:
                                        pass  # Skip if conversion fails
                        except:
                            pass  # Use original if optimization fails
                        
                        # Handle date columns
                        df = self._process_date_columns(df)
                        
                        yield df
                
                except Exception as e:
                    logger.error(f"Error processing sheet {sheet_name} in chunks: {str(e)}")
                    # Create error DataFrame
                    error_df = pd.DataFrame({
                        'sheet_name': [sheet_name],
                        'error': [str(e)]
                    })
                    yield error_df
                
                finally:
                    # Close workbook to release resources
                    workbook.close()
                    
                    # Clean up memory
                    self.memory_monitor.cleanup_memory()
                    
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                # Return error DataFrame
                error_df = pd.DataFrame({
                    'sheet_name': [sheet_name],
                    'error': [str(e)]
                })
                yield error_df
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get metadata about an Excel file without reading its full contents.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with file metadata
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
            
        file_info = {
            "file_path": file_path,
            "file_name": os.path.basename(file_path),
            "file_size_bytes": os.path.getsize(file_path),
            "file_size_mb": os.path.getsize(file_path) / (1024 * 1024),
            "file_type": "excel",
            "sheets": {}
        }
        
        try:
            # Use openpyxl to get metadata (more memory efficient than pandas for this task)
            workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
            
            # Get sheet names
            sheet_names = workbook.sheetnames
            file_info["sheet_names"] = sheet_names
            file_info["sheet_count"] = len(sheet_names)
            
            # Get sheet-specific info
            for sheet_name in sheet_names:
                try:
                    sheet = workbook[sheet_name]
                    
                    # Get basic sheet dimensions (these can be expensive operations in large files)
                    row_count = 0
                    col_count = 0
                    
                    # Sample the first few rows to determine structure
                    sample_data = []
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i >= 10:  # Only check first 10 rows
                            break
                            
                        sample_data.append(row)
                        row_count += 1
                        col_count = max(col_count, len(row))
                    
                    # Attempt to detect if first row is header
                    has_header = False
                    if row_count >= 2:
                        # Check if first row has different types than data rows
                        header_row = sample_data[0]
                        data_rows = sample_data[1:]
                        
                        # Check if header contains string labels while data contains numbers
                        header_is_text = all(isinstance(cell, str) for cell in header_row if cell is not None)
                        data_has_nums = any(any(isinstance(cell, (int, float)) for cell in row if cell is not None) 
                                         for row in data_rows)
                                         
                        has_header = header_is_text and data_has_nums
                    
                    # Estimate number of rows if possible
                    # Note: sheet.max_row can be inaccurate in read_only mode, but we can use it as a hint
                    estimated_rows = getattr(sheet, 'max_row', 0)
                    
                    # Store sheet info
                    file_info["sheets"][sheet_name] = {
                        "row_count": estimated_rows,
                        "column_count": col_count,
                        "has_header": has_header,
                        "sample_row_count": row_count
                    }
                except Exception as e:
                    logger.warning(f"Error getting info for sheet {sheet_name}: {str(e)}")
                    file_info["sheets"][sheet_name] = {
                        "error": str(e)
                    }
            
            # Close workbook to release resources
            workbook.close()
            
            # Try to read a small sample with pandas to get column info
            try:
                # Use the first sheet as sample
                if sheet_names:
                    sample_df = pd.read_excel(file_path, sheet_name=sheet_names[0], nrows=5)
                    
                    file_info["sample_columns"] = list(sample_df.columns)
                    file_info["sample_dtypes"] = {col: str(sample_df[col].dtype) for col in sample_df.columns}
            except Exception as e:
                logger.warning(f"Could not sample columns: {str(e)}")
            
            # Estimate memory requirements
            file_info["estimated_memory_mb"] = self.estimate_memory_requirement(file_path)
            
            return file_info
            
        except Exception as e:
            logger.error(f"Error getting Excel file info for {file_path}: {str(e)}")
            file_info["error"] = str(e)
            # Return basic file info even if detailed analysis fails
            return file_info
    
    def _process_date_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Process date columns in the DataFrame.
        
        Args:
            df: DataFrame to process
            
        Returns:
            DataFrame with processed date columns
        """
        # Convert specified date columns
        if self.date_columns:
            for col in self.date_columns:
                if col in df.columns:
                    try:
                        df[col] = pd.to_datetime(df[col], format=self.date_format, errors='coerce')
                    except Exception as e:
                        logger.warning(f"Error converting column '{col}' to datetime: {str(e)}")
        
        # Try to automatically detect date columns if format provided
        elif self.date_format:
            for col in df.columns:
                # Skip columns that are already datetime
                if pd.api.types.is_datetime64_dtype(df[col]):
                    continue
                    
                # Skip numeric columns
                if pd.api.types.is_numeric_dtype(df[col]):
                    continue
                
                # Check if column name contains date indicators
                if any(date_term in col.lower() for date_term in ['date', 'time', 'dt', 'day']):
                    try:
                        df[col] = pd.to_datetime(df[col], format=self.date_format, errors='coerce')
                    except:
                        pass  # Skip if conversion fails
                        
        return df
        
    def extract_sheet_names(self, file_path: str) -> List[str]:
        """
        Extract sheet names from Excel file with minimal memory usage.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of sheet names
        """
        try:
            workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            logger.error(f"Error extracting sheet names: {str(e)}")
            return []

    def sample_data(self, file_path: str, sample_size: int = 1000) -> Optional[pd.DataFrame]:
        """
        Get a sample of data from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            sample_size: Number of rows to sample
            
        Returns:
            DataFrame with sample data or None if error
        """
        try:
            # Get sheet names first
            sheet_names = self.extract_sheet_names(file_path)
            if not sheet_names:
                return None
            
            # Use the first sheet for sampling
            sheet_name = sheet_names[0]
            
            # Read a sample from the first sheet
            sample_df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                nrows=sample_size,
                header=self.header,
                skiprows=self.skiprows,
                usecols=self.usecols,
                dtype=self.dtype
            )
            
            # Add sheet name as metadata column
            sample_df['_sheet_name'] = sheet_name
            
            # Handle date columns
            sample_df = self._process_date_columns(sample_df)
            
            logger.debug(f"Sampled {len(sample_df)} rows from {os.path.basename(file_path)} (sheet: {sheet_name})")
            return sample_df
            
        except Exception as e:
            logger.error(f"Error sampling data from {file_path}: {str(e)}")
            return None

    def process_file(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """
        Process an Excel file using the provided function or default processing.
        This method implements the abstract method from BaseProcessor.
        
        Args:
            file_path: Path to the Excel file to process
            **kwargs: Additional processing parameters including:
                - process_fn: Function to apply to each chunk
                - output_path: Path to save processed output
                
        Returns:
            Dictionary with processing results
        """
        start_time = time.time()
        
        try:
            # Extract parameters
            process_fn = kwargs.get('process_fn')
            output_path = kwargs.get('output_path')
            
            # Initialize results
            total_rows_processed = 0
            processed_chunks = []
            
            # Track memory usage
            self.memory_monitor.start_tracking_step(f"excel_processing_{os.path.basename(file_path)}")
            
            # Process file in chunks
            for chunk in self.read_file(file_path):
                chunk_rows = len(chunk)
                
                # Apply processing function if provided
                if process_fn:
                    try:
                        chunk = process_fn(chunk)
                    except Exception as e:
                        logger.warning(f"Error applying process function to chunk: {str(e)}")
                
                processed_chunks.append(chunk)
                total_rows_processed += chunk_rows
                
                # Update memory peak tracking
                self.memory_monitor.update_step_peak(f"excel_processing_{os.path.basename(file_path)}")
            
            # Combine results if we have chunks to process
            if processed_chunks:
                combined_df = pd.concat(processed_chunks, ignore_index=True)
                
                # Save output if path provided
                if output_path:
                    # Determine output format based on extension
                    if output_path.endswith('.xlsx'):
                        combined_df.to_excel(output_path, index=False)
                    else:
                        combined_df.to_csv(output_path, index=False)
                    logger.info(f"Saved processed Excel to {output_path}")
            else:
                combined_df = pd.DataFrame()
            
            # End memory tracking
            memory_stats = self.memory_monitor.end_tracking_step(f"excel_processing_{os.path.basename(file_path)}")
            
            # Calculate processing time
            processing_time = time.time() - start_time
            
            return {
                "status": "completed",
                "file_path": file_path,
                "output_path": output_path,
                "data": combined_df,
                "stats": {
                    "total_rows_processed": total_rows_processed,
                    "total_chunks": len(processed_chunks),
                    "processing_time": processing_time,
                    "memory_peak": memory_stats.get("peak_mb", 0),
                    "file_size_mb": os.path.getsize(file_path) / (1024 * 1024)
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing Excel file {file_path}: {str(e)}")
            return {
                "status": "failed",
                "error": str(e),
                "file_path": file_path,
                "stats": {
                    "total_rows_processed": 0,
                    "processing_time": time.time() - start_time,
                    "memory_peak": 0
                }
            }